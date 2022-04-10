from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb_tutorial3 = load_workbook('Tutorial2.xlsx')
ws_tutorial3 = wb_tutorial3['Data']

# Looping all rows
for row in range(1, 21):
    for col in range(1, 4):
        char = get_column_letter(col)
        print(ws_tutorial3[char + str(row)].value)

# Merging Cells
ws_tutorial3.merge_cells("A1:C1")

# Inserting rows
ws_tutorial3.insert_rows(7)

# Deleting rows
ws_tutorial3.delete_rows(7)

# We can do the same with columns using insert_cols, and delete_cols

wb_tutorial3.save('Tutorial2.xlsx')