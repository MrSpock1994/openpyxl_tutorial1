from openpyxl import workbook, load_workbook

# We have to instantiate the workbook
wb_grades = load_workbook('Grades.xlsx')
print(wb_grades)

# Specifying a sheet
ws1 = wb_grades['Grades']

# Accessing a cell value, both ways are possible
print(ws1['A1'].value)
print(ws1.cell(1, 1).value)

# Changing cell value
ws1.cell(2, 1).value = "William"
# Saving the workbook
wb_grades.save('Grades.xlsx')
# Checking that the name was changed
print(ws1.cell(2, 1).value)

# Printing all the sheets name
print(wb_grades.sheetnames)

# Accessing other sheets
ws2 = wb_grades['Test1']

# Now im going to insert a value to the cell (20, 1)
ws2.cell(20, 1).value = "Testing"

# We are going to create a while loop to check and insert a value
i = 1
while ws2.cell(i, 1).value is None:
    ws2.cell(i, 1).value = "INSERTING"
    i += 1

# In the loop above where are iterating and inserting a value in all empty cells before the first non empty

# Creating a new sheet
wb_grades.create_sheet("Test3")

# Printing all the sheets name
print(wb_grades.sheetnames)

wb_grades.save('Grades.xlsx')

