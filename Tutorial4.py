from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Joe": {
        "math": 65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Bill": {
        "math": 55,
        "science": 72,
        "english": 87,
        "gym": 95
    },
    "Tim": {
        "math": 100,
        "science": 45,
        "english": 75,
        "gym": 92
    },
    "Sally": {
        "math": 30,
        "science": 25,
        "english": 45,
        "gym": 100
    },
    "Jane": {
        "math": 100,
        "science": 100,
        "english": 100,
        "gym": 60
    }
}

wb_tutorial4 = Workbook()
ws_tutorial4 = wb_tutorial4.active
ws_tutorial4.title = "Students_Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws_tutorial4.append(headings)

for person in data:
    grades = list(data[person].values())
    ws_tutorial4.append([person] + grades)

# Calculating the average
for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)
    ws_tutorial4[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
    ws_tutorial4[get_column_letter(col) + '1'].font = Font(bold=True)
wb_tutorial4.save('New_Grades.xlsx')